"""
modules/export_manager.py
Builds a pandas DataFrame from CVResult objects and provides helpers
for CSV, Excel, and PDF export (Streamlit download buttons).

Excel export: multi-sheet workbook with conditional formatting
  - Sheet 1: "Tất cả ứng viên"
  - Sheet 2: "Shortlist"
  - Sheet 3: "Thống kê JD"

PDF export: per-candidate report cards using fpdf2, bundled into one PDF.
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any

import pandas as pd
from fpdf import FPDF
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from modules.cv_analyzer import CVResult

# Matches characters that are illegal in Excel/openpyxl worksheets
# (control chars except tab, newline, carriage-return)
_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# Colour fills for score bands
_FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_FILL_RED    = PatternFill("solid", fgColor="FFC7CE")
_FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
_FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",  vertical="center", wrap_text=True)
_THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# Threshold for colour-coding
_SCORE_HIGH = 70.0
_SCORE_MED  = 50.0


def _sanitize(value: object) -> object:
    """Strip illegal Excel characters from string values."""
    if isinstance(value, str):
        return _ILLEGAL_CHARS_RE.sub("", value)
    return value


def results_to_dataframe(results: list[CVResult]) -> pd.DataFrame:
    """
    Convert a ranked list of CVResults into a tidy DataFrame.

    Columns
    -------
    Rank | Candidate Name | Email | Phone | Filename | Match Score (%)
    | Semantic (%) | Skill Coverage (%) | Experience (years) | Has Projects
    | Job Hopping | Must-have Missing | Nice-to-have Missing
    | Tags | Red Flags | Skills Found | Skills Missing | Summary | Top Evidence
    """
    rows = []
    for rank, r in enumerate(results, start=1):
        evidence_text = " | ".join(chunk for chunk, _ in r.evidence[:3])
        rows.append(
            {
                "Rank": rank,
                "Candidate Name": r.candidate_name or "—",
                "Email": r.email or "—",
                "Phone": r.phone or "—",
                "Filename": r.filename,
                "Match Score (%)": round(r.score * 100, 1),
                "Semantic (%)": round(r.semantic_score * 100, 1),
                "Skill Coverage (%)": round(r.skill_score * 100, 1),
                "Experience (years)": r.experience_years if r.experience_years else "—",
                "Has Projects": "Yes" if r.has_projects else "No",
                "Job Hopping": "Yes" if r.job_hopping else "No",
                "Must-have Missing": ", ".join(r.must_have_missing) if r.must_have_missing else "—",
                "Nice-to-have Missing": ", ".join(r.nice_to_have_missing) if r.nice_to_have_missing else "—",
                "Tags": " | ".join(r.tags) if r.tags else "—",
                "Red Flags": " | ".join(r.red_flags) if r.red_flags else "—",
                "Skills Found": ", ".join(r.skills_found) if r.skills_found else "—",
                "Skills Missing": ", ".join(r.skills_missing) if r.skills_missing else "—",
                "Summary": r.summary,
                "Top Evidence": evidence_text,
            }
        )
    return pd.DataFrame(rows)


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Return the DataFrame encoded as UTF-8 CSV bytes."""
    return df.to_csv(index=False).encode("utf-8")


def _style_sheet(ws: Any, df: pd.DataFrame, score_col_idx: int) -> None:
    """Apply header styling and conditional formatting to an openpyxl worksheet."""
    # Header row
    for cell in ws[1]:
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _THIN_BORDER

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-width columns (capped at 60)
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)), *(len(str(v)) for v in df.iloc[:, col_idx - 1]))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    # Conditional fill on score column and data rows
    for row_idx in range(2, ws.max_row + 1):
        score_cell = ws.cell(row=row_idx, column=score_col_idx)
        try:
            score_val = float(score_cell.value)
        except (TypeError, ValueError):
            score_val = -1

        fill = (
            _FILL_GREEN  if score_val >= _SCORE_HIGH else
            _FILL_YELLOW if score_val >= _SCORE_MED  else
            _FILL_RED    if score_val >= 0            else
            None
        )

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _THIN_BORDER
            cell.alignment = _ALIGN_LEFT
            if fill and col_idx == score_col_idx:
                cell.fill = fill


def to_excel_bytes(df: pd.DataFrame, jd_skills: list[str] | None = None) -> bytes:
    """
    Return a multi-sheet Excel workbook as bytes.

    Sheets
    ------
    1. Tất cả ứng viên  — full ranked list with styling
    2. Shortlist        — candidates with score >= 70%
    3. Thống kê JD      — skill gap summary table
    """
    jd_skills = jd_skills or []
    clean_df = df.apply(lambda col: col.map(_sanitize))

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # --- Sheet 1: All candidates ---
        clean_df.to_excel(writer, index=False, sheet_name="Tất cả ứng viên")
        ws_all = writer.sheets["Tất cả ứng viên"]
        score_col = list(clean_df.columns).index("Match Score (%)") + 1
        _style_sheet(ws_all, clean_df, score_col)

        # --- Sheet 2: Shortlist ---
        shortlist_df = clean_df[clean_df["Match Score (%)"] >= _SCORE_HIGH].copy()
        if not shortlist_df.empty:
            shortlist_df.to_excel(writer, index=False, sheet_name="Shortlist")
            ws_sl = writer.sheets["Shortlist"]
            _style_sheet(ws_sl, shortlist_df, score_col)
        else:
            pd.DataFrame({"Info": ["Chưa có ứng viên đạt ≥70%"]}).to_excel(
                writer, index=False, sheet_name="Shortlist"
            )

        # --- Sheet 3: JD Skill Gap ---
        if jd_skills and "Skills Found" in clean_df.columns:
            skill_rows = []
            total = len(clean_df)
            for skill in jd_skills:
                found_count = sum(
                    1 for val in clean_df["Skills Found"]
                    if isinstance(val, str) and skill in val
                )
                missing_count = total - found_count
                skill_rows.append({
                    "Kỹ năng JD": skill,
                    "Số CV có": found_count,
                    "Số CV thiếu": missing_count,
                    "Tỷ lệ có (%)": round(found_count / total * 100, 1) if total else 0,
                    "Gợi ý": "⚠️ Thiếu nhiều" if found_count / max(total, 1) < 0.4 else "✅ OK",
                })
            stats_df = pd.DataFrame(skill_rows)
            stats_df.to_excel(writer, index=False, sheet_name="Thống kê JD")
            ws_stats = writer.sheets["Thống kê JD"]
            for cell in ws_stats[1]:
                cell.fill = _FILL_HEADER
                cell.font = _FONT_HEADER
                cell.alignment = _ALIGN_CENTER
            ws_stats.freeze_panes = "A2"
            for col_idx, col_name in enumerate(stats_df.columns, start=1):
                ws_stats.column_dimensions[get_column_letter(col_idx)].width = max(
                    len(str(col_name)) + 4, 18
                )

    return buffer.getvalue()


# ---------------------------------------------------------------------------
# PDF batch report (fpdf2)
# ---------------------------------------------------------------------------

def _safe_str(value: object, max_len: int = 200) -> str:
    """Convert value to a safe ASCII-ish string for fpdf (latin-1 safe)."""
    text = str(value) if value is not None else ""
    # Replace characters outside latin-1 range with their closest ASCII equivalent
    text = text.encode("latin-1", errors="replace").decode("latin-1")
    return text[:max_len]


def to_pdf_bytes(results: list[CVResult], jd_snippet: str = "") -> bytes:
    """
    Generate a multi-page PDF report for a list of CVResults.

    One summary page (all candidates ranked) + one detail page per candidate.

    Parameters
    ----------
    results     : list[CVResult]  ranked candidate results
    jd_snippet  : str             short excerpt of the job description

    Returns
    -------
    bytes  – raw PDF bytes suitable for st.download_button
    """
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(15, 15, 15)

    # ------------------------------------------------------------------ #
    # Cover / summary page                                                 #
    # ------------------------------------------------------------------ #
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, "CV Screening Report", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 7, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True, align="C")
    pdf.ln(4)

    if jd_snippet:
        pdf.set_font("Helvetica", "I", 9)
        pdf.multi_cell(0, 6, f"JD: {_safe_str(jd_snippet, 160)}")
        pdf.ln(4)

    # Summary table header
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(31, 78, 121)   # dark blue
    pdf.set_text_color(255, 255, 255)
    col_widths = [12, 55, 22, 22, 22, 22, 25]
    headers = ["Rank", "Candidate / File", "Score%", "Sem%", "Skill%", "Exp(yr)", "Decision"]
    for w, h in zip(col_widths, headers):
        pdf.cell(w, 8, h, border=1, fill=True, align="C")
    pdf.ln()

    # Summary table rows
    pdf.set_text_color(0, 0, 0)
    for rank, r in enumerate(results, start=1):
        pdf.set_font("Helvetica", "", 8)
        score_pct = round(r.score * 100, 1)
        # Row colour: green / yellow / red
        if score_pct >= 70:
            pdf.set_fill_color(198, 239, 206)
        elif score_pct >= 50:
            pdf.set_fill_color(255, 235, 156)
        else:
            pdf.set_fill_color(255, 199, 206)

        name_str = _safe_str(r.candidate_name or r.filename, 40)
        values = [
            str(rank),
            name_str,
            f"{score_pct}",
            f"{round(r.semantic_score * 100, 1)}",
            f"{round(r.skill_score * 100, 1)}",
            str(r.experience_years) if r.experience_years else "-",
            "-",
        ]
        for w, v in zip(col_widths, values):
            pdf.cell(w, 7, v, border=1, fill=True, align="C")
        pdf.ln()

    # ------------------------------------------------------------------ #
    # Detail pages — one per candidate                                     #
    # ------------------------------------------------------------------ #
    for rank, r in enumerate(results, start=1):
        pdf.add_page()

        # Header bar
        pdf.set_fill_color(31, 78, 121)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 13)
        score_pct = round(r.score * 100, 1)
        name_display = _safe_str(r.candidate_name or r.filename, 60)
        pdf.cell(0, 10, f"#{rank}  {name_display}  -  Score: {score_pct}%",
                 border=0, fill=True, ln=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(3)

        # Contact info
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(35, 6, "Email:", border=0)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 6, _safe_str(r.email or "-"), ln=True)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(35, 6, "Phone:", border=0)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 6, _safe_str(r.phone or "-"), ln=True)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(35, 6, "File:", border=0)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 6, _safe_str(r.filename, 80), ln=True)
        pdf.ln(3)

        # Score breakdown
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(220, 230, 241)
        pdf.cell(0, 7, "Score Breakdown", border=1, fill=True, ln=True)
        pdf.set_font("Helvetica", "", 9)
        breakdown = [
            ("Composite Score", f"{score_pct}%"),
            ("Semantic", f"{round(r.semantic_score * 100, 1)}%"),
            ("Skill Coverage", f"{round(r.skill_score * 100, 1)}%"),
            ("Experience", f"{r.experience_years} yr" if r.experience_years else "-"),
        ]
        for label, val in breakdown:
            pdf.cell(60, 6, label, border="LR")
            pdf.cell(0, 6, val, border="R", ln=True)
        pdf.ln(3)

        # Skills found
        if r.skills_found:
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_fill_color(220, 230, 241)
            pdf.cell(0, 7, "Skills Found", border=1, fill=True, ln=True)
            pdf.set_font("Helvetica", "", 9)
            pdf.multi_cell(0, 6, _safe_str(", ".join(r.skills_found), 300))
            pdf.ln(2)

        # Skills missing
        if r.skills_missing:
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_fill_color(255, 235, 156)
            pdf.cell(0, 7, "Skills Missing", border=1, fill=True, ln=True)
            pdf.set_font("Helvetica", "", 9)
            pdf.multi_cell(0, 6, _safe_str(", ".join(r.skills_missing), 300))
            pdf.ln(2)

        # Must-have missing (red)
        if r.must_have_missing:
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_fill_color(255, 199, 206)
            pdf.cell(0, 7, "Must-Have Missing (!)", border=1, fill=True, ln=True)
            pdf.set_font("Helvetica", "", 9)
            pdf.multi_cell(0, 6, _safe_str(", ".join(r.must_have_missing), 200))
            pdf.ln(2)

        # Tags
        if r.tags:
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_fill_color(220, 230, 241)
            pdf.cell(0, 7, "Tags", border=1, fill=True, ln=True)
            pdf.set_font("Helvetica", "", 9)
            pdf.multi_cell(0, 6, _safe_str(" | ".join(r.tags), 200))
            pdf.ln(2)

        # Red flags
        if r.red_flags:
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_fill_color(255, 199, 206)
            pdf.cell(0, 7, "Red Flags", border=1, fill=True, ln=True)
            pdf.set_font("Helvetica", "", 9)
            for flag in r.red_flags:
                pdf.multi_cell(0, 6, f"  - {_safe_str(flag, 120)}")
            pdf.ln(2)

        # Summary
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(220, 230, 241)
        pdf.cell(0, 7, "AI Summary", border=1, fill=True, ln=True)
        pdf.set_font("Helvetica", "", 8)
        pdf.multi_cell(0, 5, _safe_str(r.summary, 500))

    return bytes(pdf.output())
